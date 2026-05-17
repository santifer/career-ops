import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Scores mapeados das summaries dos scans
scores = {
    # Scan 1 — added_linear
    "https://www.databricks.com/company/careers/professional-services-operations/ai-engineer---fde-forward-deployed-engineer-8189900002": 10.0,
    "https://jobs.elastic.co/form?gh_jid=7607148": 9.0,
    "https://vercel.com/careers/ai-engineer-5517523004": 8.5,
    "https://careers.salesforce.com/en/jobs/jr305674/ai-engineer-forward-deployed-engineer-multiple-levels/": 8.0,
    "https://careers.salesforce.com/en/jobs/jr305671/ai-architect-forward-deployed-engineer-deployment-strategist-multiple-levels/": 8.0,
    "https://jobs.ashbyhq.com/lovable/9f4963e7-be14-4dd9-99ce-05df2f06e22d": 8.0,
    "https://job-boards.greenhouse.io/contentful/jobs/7487850": 8.0,
    "https://job-boards.greenhouse.io/contentful/jobs/7776562": 8.0,
    "https://job-boards.greenhouse.io/monzo/jobs/7118972": 8.0,
    "https://job-boards.greenhouse.io/monzo/jobs/7613712": 8.0,
    "https://jobs.sap.com/job/Walldorf-(Senior)-(agentic)-AI-Engineer-at-AI-CoE-(fmd)-69190/1248020901/": 7.5,
    "https://www.revolut.com/en-US/careers/position/data-scientist-nlp-deep-learning-engineer-7fdeec15-cd49-4ca9-b509-ae6ca2613cd5/": 7.5,
    # Scan 1 — pipeline
    "https://careers.salesforce.com/en/jobs/jr305198/ai-forward-deployed-engineer-seniorleadprincipal/": 7.5,
    "https://jobs.sap.com/job/Garching-bei-M%C3%BCnchen-Lead-AI-Developer-(fmd)-85748/1371474133/": 7.0,
    "https://jobs.sap.com/job/Walldorf-(Senior)-AI-Architect-(fmd)-69190/1375529133/": 7.0,
    "https://jobs.ashbyhq.com/legora/f3c0712a-f8e2-4dc1-8e83-23da7891a1c2": 7.0,
    "https://job-boards.greenhouse.io/adyen/jobs/7434460": 7.0,
    "https://careers.adyen.com/vacancies/5022994-senior-machine-learning-scientist-gen-ai/": 7.0,
    "https://job-boards.greenhouse.io/adyen/jobs/7068990": 7.0,
    "https://www.google.com/about/careers/applications/jobs/results/132616772468515526-machine-learning-engineer/": 7.0,
    "https://jobs.sap.com/job/Walldorf-Machine-Learning-Engineer-(Associate)-Developer-(fmd)-69190/1293299301/": 6.5,
    "https://www.celonis.com/careers/jobs/details/?jobId=5798566003": 7.0,
    "https://wise.jobs/jobs": 7.0,
    "https://hitmarker.net/jobs/nvidia-deep-learning-software-engineer-llm-performance-new-college-grad-2026-1685942": 7.0,
    "https://jobs.ashbyhq.com/lovable/7fe39289-1f7f-47d4-8002-d3aeeaaaabc6": 7.0,
    "https://www.amazon.jobs/content/en/artificial-intelligence-ai": 7.0,
    "https://jobs.joinimagine.com/companies/celonis/jobs/44075821-senior-software-engineer-python-ai-ml-infrastructure": 6.5,
    "https://job-boards.greenhouse.io/monzo/jobs/7766052": 6.5,
    "https://job-boards.greenhouse.io/monzo/jobs/7686352": 6.5,
    "https://jobs.ashbyhq.com/legora/3fcdc6ba-e35f-470e-823e-7c4563c933a2": 6.0,
    "https://jobs.ashbyhq.com/lovable/e27e931e-79f3-483a-b543-57e42633ac5c": 6.0,
    "https://jaabz.com/jobs/91538-senior-software-engineer-machine-learning-platform-all-genders": 6.0,
    "https://traderepublic.com/en-de/about?jobId=6302030003": 6.0,
    "https://www.typeform.com/careers/engineering": 6.0,
    "https://job-boards.greenhouse.io/contentful/jobs/7544099": 6.0,
    "https://jobs.microsoft.com": 5.0,
    "https://databricks.com/company/careers/open-positions/job?gh_jid=8341837002": 6.5,
    "https://databricks.com/company/careers/open-positions/job?gh_jid=8285292002": 6.5,
    "https://databricks.com/company/careers/open-positions/job?gh_jid=8486738002": 6.5,
    "https://databricks.com/company/careers/open-positions/job?gh_jid=8425303002": 6.5,
    # Scan 2 — AI Labs added_linear
    "https://job-boards.greenhouse.io/anthropic/jobs/4985877008": 9.0,
    "https://job-boards.greenhouse.io/anthropic/jobs/5012991008": 9.0,
    "https://job-boards.greenhouse.io/anthropic/jobs/5116274008": 8.0,
    "https://job-boards.greenhouse.io/anthropic/jobs/5068226008": 8.0,
    "https://job-boards.greenhouse.io/anthropic/jobs/5159608008": 8.0,
    "https://jobs.ashbyhq.com/openai/0c5d1302-0e77-4ff6-a2b3-c7d7bac31833": 8.0,
    "https://jobs.ashbyhq.com/openai/28cd6fe2-4096-426b-8b08-52a66458e2c6": 8.0,
    "https://jobs.ashbyhq.com/openai/941bad28-7abe-43c7-b20a-2bc7e5b3c6e8": 7.5,
    "https://jobs.lever.co/mistral/77f6fd1b-65cf-45d8-9b68-594c62732f62": 9.0,
    "https://jobs.lever.co/mistral/b56f6523-b7d2-47a3-abe6-2c793c7672fc": 9.0,
    "https://jobs.lever.co/mistral/042d7b29-279b-48e2-a44b-c7bdc3180dab": 9.0,
    "https://jobs.lever.co/mistral/07447e1d-7900-46d4-b61b-186f2f76847f": 8.0,
    "https://jobs.lever.co/mistral/aceffeba-c4e9-4b3b-adff-e7e78b986c5c": 8.0,
    "https://jobs.ashbyhq.com/cohere/b0bcef37-1d20-414f-aade-c54942d63df9": 9.0,
    "https://jobs.ashbyhq.com/cohere/38f75a48-199c-4325-a7f8-2af6ed6a1b3b": 9.0,
    "https://jobs.ashbyhq.com/cohere/1fa01a03-9253-4f62-8f10-0fe368b38cb9": 8.0,
    "https://jobs.ashbyhq.com/cohere/f931c326-4d16-44b3-812c-99378c60280f": 8.0,
    "https://jobs.ashbyhq.com/cohere/554a9380-ab50-4338-88a9-c6b8ab19d92e": 8.0,
    # Scan 2 — AI Labs pipeline
    "https://job-boards.greenhouse.io/anthropic/jobs/5121563008": 8.0,
    "https://job-boards.greenhouse.io/anthropic/jobs/5121561008": 8.0,
    "https://job-boards.greenhouse.io/anthropic/jobs/5057647008": 8.0,
    "https://job-boards.greenhouse.io/anthropic/jobs/5073277008": 8.0,
    "https://job-boards.greenhouse.io/anthropic/jobs/5111942008": 8.0,
    "https://job-boards.greenhouse.io/anthropic/jobs/5014500008": 8.0,
    "https://job-boards.greenhouse.io/anthropic/jobs/5055488008": 8.0,
    "https://job-boards.greenhouse.io/anthropic/jobs/5156057008": 7.5,
    "https://jobs.lever.co/mistral/675b7f06-a76b-4144-af0c-4dd3282ef489": 8.0,
    "https://jobs.lever.co/mistral/b7ae8fc4-5779-4ad2-8f5b-632b4d9498cf": 8.0,
    "https://jobs.lever.co/mistral/c79ff8ed-6689-4dda-aec6-979a5dc767d0": 7.5,
    "https://jobs.lever.co/mistral/db67d7a2-bcec-4151-9b3a-8212ddabf419": 7.0,
    "https://jobs.ashbyhq.com/openai/b9dee2a0-9bb3-447e-9bce-2b1bed784e5b": 7.0,
    "https://jobs.ashbyhq.com/openai/d44c9f70-4aef-45a4-a36a-54fb65663ccb": 7.0,
    "https://jobs.ashbyhq.com/openai/5f0c6579-0bfb-4a06-8a43-1dd371499e10": 7.0,
    "https://jobs.ashbyhq.com/openai/7055f010-99f4-4c76-8361-ba5b5f9af1d0": 7.0,
    "https://jobs.ashbyhq.com/openai/7322d344-9325-4a92-8445-0a2c4e9272f8": 6.0,
    "https://jobs.ashbyhq.com/openai/d8794980-1d3f-4d82-8b48-811449b6c492": 6.0,
    "https://jobs.ashbyhq.com/cohere/e912d84c-8399-422d-8a7d-918422a3e4b1": 8.0,
    "https://jobs.ashbyhq.com/cohere/f5f727bc-9eab-4fcc-9fdd-92a9ce47c37c": 8.0,
    "https://jobs.ashbyhq.com/cohere/110ba167-4efd-43b7-85d2-3ff719a28b0f": 8.0,
    "https://jobs.ashbyhq.com/cohere/24fe6a0b-6209-4ee0-b622-49c18636d99c": 8.0,
}

# Ler TSV
rows = []
with open("/Users/marlow/career-ops/data/scan-history.tsv", encoding="utf-8") as f:
    lines = f.readlines()

header = lines[0].strip().split("\t")
for line in lines[1:]:
    parts = line.strip().split("\t")
    if len(parts) < 6:
        continue
    url, first_seen, portal, title, company, status = parts[0], parts[1], parts[2], parts[3], parts[4], parts[5]
    if status not in ("added_linear", "pipeline"):
        continue
    score = scores.get(url, "")
    rows.append({
        "Empresa": company,
        "Vaga": title,
        "Score": score,
        "Linear": "✅" if status == "added_linear" else "—",
        "Portal": portal,
        "Data": first_seen,
        "URL": url,
    })

# Ordenar por score desc
rows.sort(key=lambda r: (-(r["Score"] if r["Score"] != "" else 0)))

# Criar workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Prospect Scan 2026-04-11"

# Estilos
header_fill = PatternFill("solid", fgColor="1E293B")
header_font = Font(bold=True, color="FFFFFF", size=11)
linear_fill = PatternFill("solid", fgColor="D1FAE5")  # verde claro
pipeline_fill = PatternFill("solid", fgColor="FEF3C7")  # amarelo claro

score_colors = {
    (9.0, 10.0): "065F46",  # verde escuro
    (7.5, 8.9): "047857",   # verde médio
    (5.0, 7.4): "92400E",   # âmbar
}

thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header
cols = ["Empresa", "Vaga", "Score", "Linear", "Portal", "Data", "URL"]
col_widths = [18, 45, 8, 8, 14, 12, 60]

for i, col in enumerate(cols, 1):
    cell = ws.cell(row=1, column=i, value=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border
    ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]

ws.row_dimensions[1].height = 22

# Dados
for r_idx, row in enumerate(rows, 2):
    is_linear = row["Linear"] == "✅"
    row_fill = linear_fill if is_linear else pipeline_fill

    for c_idx, col in enumerate(cols, 1):
        val = row[col]
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.fill = row_fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(col == "Vaga"))

        if col == "Score" and val != "":
            cell.alignment = Alignment(horizontal="center", vertical="center")
            color = "374151"  # cinza padrão
            for (lo, hi), c in score_colors.items():
                if lo <= val <= hi:
                    color = c
                    break
            cell.font = Font(bold=True, color=color)

        if col == "URL":
            cell.hyperlink = val
            cell.font = Font(color="2563EB", underline="single")

    ws.row_dimensions[r_idx].height = 18

# Freeze header
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:G{len(rows)+1}"

# Legenda na aba lateral
ws2 = wb.create_sheet("Legenda")
ws2["A1"] = "Cor"
ws2["B1"] = "Significado"
ws2.cell(1,1).font = Font(bold=True)
ws2.cell(1,2).font = Font(bold=True)

legenda = [
    ("Verde (linha)", "Score ≥ 7.5 → Issue criada no Linear"),
    ("Amarelo (linha)", "Score 5.0–7.4 → No pipeline para avaliação posterior"),
    ("Score verde escuro", "Score 9–10: aplicar imediatamente"),
    ("Score verde médio", "Score 7.5–8.9: forte match"),
    ("Score âmbar", "Score 5.0–7.4: match razoável"),
]
for i, (cor, sig) in enumerate(legenda, 2):
    ws2.cell(i, 1, cor)
    ws2.cell(i, 2, sig)

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 45

out = "/Users/marlow/career-ops/output/prospect-scan-2026-04-11.xlsx"
import os; os.makedirs("/Users/marlow/career-ops/output", exist_ok=True)
wb.save(out)
print(f"Salvo: {out} ({len(rows)} vagas)")
